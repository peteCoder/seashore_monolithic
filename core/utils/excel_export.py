"""
Excel/CSV Export Utilities using Pandas
========================================

Professional Excel exports for accounting reports
"""

from django.http import HttpResponse
import pandas as pd
from io import BytesIO
from datetime import datetime


def create_excel_response(filename='report.xlsx'):
    """Create an HTTP response for Excel file download"""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def create_csv_response(filename='report.csv'):
    """Create an HTTP response for CSV file download"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_trial_balance_excel(report_data, form_data):
    """Export Trial Balance to Excel"""
    # Create Excel writer
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')

    # Prepare data
    trial_balance_data = []
    for item in report_data['trial_balance']:
        trial_balance_data.append({
            'GL Code': item['account'].gl_code,
            'Account Name': item['account'].account_name,
            'Account Type': item['account'].account_type.get_name_display(),
            'Debit (₦)': float(item['debit']) if item['debit'] > 0 else 0,
            'Credit (₦)': float(item['credit']) if item['credit'] > 0 else 0,
        })

    # Create DataFrame
    df = pd.DataFrame(trial_balance_data)

    # Add totals row
    totals_row = pd.DataFrame([{
        'GL Code': '',
        'Account Name': 'TOTAL',
        'Account Type': '',
        'Debit (₦)': float(report_data['total_debits']),
        'Credit (₦)': float(report_data['total_credits']),
    }])
    df = pd.concat([df, totals_row], ignore_index=True)

    # Write to Excel
    df.to_excel(writer, sheet_name='Trial Balance', index=False)

    # Get workbook and worksheet
    workbook = writer.book
    worksheet = writer.sheets['Trial Balance']

    # Apply styling
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Header styling
    header_fill = PatternFill(start_color='D97706', end_color='D97706', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)

    for col_num, col in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Totals row styling
    totals_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
    totals_font = Font(bold=True, size=11)
    last_row = len(df) + 1

    for col_num in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=last_row, column=col_num)
        cell.fill = totals_fill
        cell.font = totals_font

    # Number formatting for currency columns
    for row in range(2, last_row + 1):
        worksheet.cell(row=row, column=4).number_format = '#,##0.00'  # Debit
        worksheet.cell(row=row, column=5).number_format = '#,##0.00'  # Credit

    # Adjust column widths
    worksheet.column_dimensions['A'].width = 12  # GL Code
    worksheet.column_dimensions['B'].width = 40  # Account Name
    worksheet.column_dimensions['C'].width = 20  # Account Type
    worksheet.column_dimensions['D'].width = 18  # Debit
    worksheet.column_dimensions['E'].width = 18  # Credit

    # Add report header information
    worksheet.insert_rows(1, 3)
    worksheet.merge_cells('A1:E1')
    worksheet.merge_cells('A2:E2')
    worksheet.merge_cells('A3:E3')

    title_cell = worksheet['A1']
    title_cell.value = 'TRIAL BALANCE'
    title_cell.font = Font(bold=True, size=16, color='D97706')
    title_cell.alignment = Alignment(horizontal='center')

    period_cell = worksheet['A2']
    period_cell.value = f'Period: {report_data["date_from"].strftime("%B %d, %Y")} to {report_data["date_to"].strftime("%B %d, %Y")}'
    period_cell.alignment = Alignment(horizontal='center')

    balance_cell = worksheet['A3']
    balance_status = 'BALANCED ✓' if report_data['is_balanced'] else 'NOT BALANCED ✗'
    balance_cell.value = f'Status: {balance_status}'
    balance_cell.font = Font(bold=True, color='059669' if report_data['is_balanced'] else 'DC2626')
    balance_cell.alignment = Alignment(horizontal='center')

    # Save
    writer.close()
    output.seek(0)

    filename = f'trial_balance_{report_data["date_from"].strftime("%Y%m%d")}_{report_data["date_to"].strftime("%Y%m%d")}.xlsx'
    response = create_excel_response(filename)
    response.write(output.read())
    return response


def export_profit_loss_excel(report_data, form_data):
    """Export Profit & Loss to Excel"""
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')

    # Income section
    income_data = []
    for item in report_data['income_items']:
        income_data.append({
            'GL Code': item['account'].gl_code,
            'Account': item['account'].account_name,
            'Amount (₦)': float(item['amount']),
        })

    # Expense section
    expense_data = []
    for item in report_data['expense_items']:
        expense_data.append({
            'GL Code': item['account'].gl_code,
            'Account': item['account'].account_name,
            'Amount (₦)': float(item['amount']),
        })

    # Create DataFrames
    df_income = pd.DataFrame(income_data)
    df_expense = pd.DataFrame(expense_data)

    # Write to separate sheets
    df_income.to_excel(writer, sheet_name='Income', index=False)
    df_expense.to_excel(writer, sheet_name='Expenses', index=False)

    # Create summary sheet
    summary_data = pd.DataFrame([
        {'Item': 'Total Income', 'Amount (₦)': float(report_data['total_income'])},
        {'Item': 'Total Expenses', 'Amount (₦)': float(report_data['total_expenses'])},
        {'Item': 'Net Profit/Loss', 'Amount (₦)': float(report_data['net_profit'])},
    ])
    summary_data.to_excel(writer, sheet_name='Summary', index=False)

    # Apply styling (similar to trial balance)
    from openpyxl.styles import Font, PatternFill, Alignment

    for sheet_name in writer.sheets:
        worksheet = writer.sheets[sheet_name]

        # Header styling
        for col_num in range(1, 4):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = PatternFill(start_color='D97706', end_color='D97706', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center')

        # Column widths
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 40
        worksheet.column_dimensions['C'].width = 18

    writer.close()
    output.seek(0)

    filename = f'profit_loss_{report_data["date_from"].strftime("%Y%m%d")}_{report_data["date_to"].strftime("%Y%m%d")}.xlsx'
    response = create_excel_response(filename)
    response.write(output.read())
    return response


def export_general_ledger_excel(report_data, form_data):
    """Export General Ledger to Excel"""
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')

    # Prepare transaction data
    transactions_data = []
    for txn in report_data['transactions']:
        transactions_data.append({
            'Date': txn['line'].journal_entry.transaction_date.strftime('%Y-%m-%d'),
            'Journal Number': txn['line'].journal_entry.journal_number,
            'Description': txn['line'].description,
            'Debit (₦)': float(txn['line'].debit_amount) if txn['line'].debit_amount > 0 else 0,
            'Credit (₦)': float(txn['line'].credit_amount) if txn['line'].credit_amount > 0 else 0,
            'Balance (₦)': float(txn['running_balance']),
        })

    df = pd.DataFrame(transactions_data)
    df.to_excel(writer, sheet_name='General Ledger', index=False)

    # Styling
    worksheet = writer.sheets['General Ledger']
    from openpyxl.styles import Font, PatternFill, Alignment

    # Add header info
    worksheet.insert_rows(1, 4)
    worksheet.merge_cells('A1:F1')
    worksheet.merge_cells('A2:F2')
    worksheet.merge_cells('A3:F3')

    worksheet['A1'] = 'GENERAL LEDGER'
    worksheet['A1'].font = Font(bold=True, size=16, color='D97706')
    worksheet['A1'].alignment = Alignment(horizontal='center')

    worksheet['A2'] = f'{report_data["account"].gl_code} - {report_data["account"].account_name}'
    worksheet['A2'].font = Font(bold=True, size=12)
    worksheet['A2'].alignment = Alignment(horizontal='center')

    worksheet['A3'] = f'Period: {report_data["date_from"].strftime("%B %d, %Y")} to {report_data["date_to"].strftime("%B %d, %Y")}'
    worksheet['A3'].alignment = Alignment(horizontal='center')

    worksheet['A4'] = f'Opening Balance: ₦{report_data["opening_balance"]:,.2f}'
    worksheet['A4'].font = Font(bold=True)

    writer.close()
    output.seek(0)

    filename = f'general_ledger_{report_data["account"].gl_code}_{report_data["date_from"].strftime("%Y%m%d")}.xlsx'
    response = create_excel_response(filename)
    response.write(output.read())
    return response


def export_balance_sheet_excel(report_data, form_data):
    """Export Balance Sheet to Excel"""
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')

    # Assets section
    assets_data = []
    for item in report_data['assets']:
        assets_data.append({
            'GL Code': item['account'].gl_code,
            'Account': item['account'].account_name,
            'Balance (₦)': float(item['balance']),
        })

    # Liabilities section
    liabilities_data = []
    for item in report_data['liabilities']:
        liabilities_data.append({
            'GL Code': item['account'].gl_code,
            'Account': item['account'].account_name,
            'Balance (₦)': float(item['balance']),
        })

    # Equity section
    equity_data = []
    for item in report_data['equity']:
        equity_data.append({
            'GL Code': item['account'].gl_code,
            'Account': item['account'].account_name,
            'Balance (₦)': float(item['balance']),
        })

    # Create DataFrames
    df_assets = pd.DataFrame(assets_data)
    df_liabilities = pd.DataFrame(liabilities_data)
    df_equity = pd.DataFrame(equity_data)

    # Write to separate sheets
    df_assets.to_excel(writer, sheet_name='Assets', index=False)
    df_liabilities.to_excel(writer, sheet_name='Liabilities', index=False)
    df_equity.to_excel(writer, sheet_name='Equity', index=False)

    # Create summary sheet
    summary_data = pd.DataFrame([
        {'Category': 'Total Assets', 'Amount (₦)': float(report_data['total_assets'])},
        {'Category': 'Total Liabilities', 'Amount (₦)': float(report_data['total_liabilities'])},
        {'Category': 'Total Equity', 'Amount (₦)': float(report_data['total_equity'])},
        {'Category': 'Total Liabilities + Equity', 'Amount (₦)': float(report_data['total_liabilities_equity'])},
    ])
    summary_data.to_excel(writer, sheet_name='Summary', index=False)

    # Apply styling
    from openpyxl.styles import Font, PatternFill, Alignment

    for sheet_name in writer.sheets:
        worksheet = writer.sheets[sheet_name]

        # Header styling
        for col_num in range(1, 4):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = PatternFill(start_color='D97706', end_color='D97706', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center')

        # Column widths
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 40
        worksheet.column_dimensions['C'].width = 18

    writer.close()
    output.seek(0)

    filename = f'balance_sheet_{report_data["as_of_date"].strftime("%Y%m%d")}.xlsx'
    response = create_excel_response(filename)
    response.write(output.read())
    return response


def export_cash_flow_excel(report_data, form_data):
    """Export Cash Flow Statement to Excel"""
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')

    # Operating activities
    operating_data = []
    for item in report_data['operating_activities']:
        operating_data.append({
            'Date': item['line'].journal_entry.transaction_date.strftime('%Y-%m-%d'),
            'Description': item['line'].description,
            'Amount (₦)': float(item['amount']),
        })

    # Investing activities
    investing_data = []
    for item in report_data['investing_activities']:
        investing_data.append({
            'Date': item['line'].journal_entry.transaction_date.strftime('%Y-%m-%d'),
            'Description': item['line'].description,
            'Amount (₦)': float(item['amount']),
        })

    # Create DataFrames
    df_operating = pd.DataFrame(operating_data) if operating_data else pd.DataFrame(columns=['Date', 'Description', 'Amount (₦)'])
    df_investing = pd.DataFrame(investing_data) if investing_data else pd.DataFrame(columns=['Date', 'Description', 'Amount (₦)'])

    # Write to separate sheets
    df_operating.to_excel(writer, sheet_name='Operating Activities', index=False)
    df_investing.to_excel(writer, sheet_name='Investing Activities', index=False)

    # Create summary sheet
    summary_data = pd.DataFrame([
        {'Activity Type': 'Operating Activities', 'Total (₦)': float(report_data['operating_total'])},
        {'Activity Type': 'Investing Activities', 'Total (₦)': float(report_data['investing_total'])},
        {'Activity Type': 'Financing Activities', 'Total (₦)': float(report_data['financing_total'])},
        {'Activity Type': 'Net Cash Flow', 'Total (₦)': float(report_data['net_cash_flow'])},
    ])
    summary_data.to_excel(writer, sheet_name='Summary', index=False)

    # Apply styling
    from openpyxl.styles import Font, PatternFill, Alignment

    for sheet_name in writer.sheets:
        worksheet = writer.sheets[sheet_name]

        # Header styling
        for col_num in range(1, 4):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = PatternFill(start_color='D97706', end_color='D97706', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center')

        # Column widths
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 50
        worksheet.column_dimensions['C'].width = 18

    writer.close()
    output.seek(0)

    filename = f'cash_flow_{report_data["date_from"].strftime("%Y%m%d")}_{report_data["date_to"].strftime("%Y%m%d")}.xlsx'
    response = create_excel_response(filename)
    response.write(output.read())
    return response


def export_transaction_audit_excel(report_data, form_data):
    """Export Transaction Audit Log to Excel"""
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')

    # Prepare audit data
    audit_data_list = []
    for item in report_data['audit_data']:
        txn = item['transaction']
        journal_status = 'Yes' if item['has_journal'] else 'NO - MISSING ⚠️'

        audit_data_list.append({
            'Date': txn.transaction_date.strftime('%Y-%m-%d'),
            'Transaction Ref': txn.transaction_ref,
            'Type': txn.transaction_type,
            'Client': txn.client.get_full_name() if txn.client else 'N/A',
            'Amount (₦)': float(txn.amount),
            'Branch': txn.branch.name if txn.branch else 'N/A',
            'Has Journal Entry': journal_status,
        })

    df = pd.DataFrame(audit_data_list)
    df.to_excel(writer, sheet_name='Audit Log', index=False)

    # Styling
    worksheet = writer.sheets['Audit Log']
    from openpyxl.styles import Font, PatternFill, Alignment

    # Add header info
    worksheet.insert_rows(1, 3)
    worksheet.merge_cells('A1:G1')
    worksheet.merge_cells('A2:G2')

    worksheet['A1'] = 'TRANSACTION AUDIT LOG'
    worksheet['A1'].font = Font(bold=True, size=16, color='D97706')
    worksheet['A1'].alignment = Alignment(horizontal='center')

    worksheet['A2'] = f'Total Transactions: {report_data["total_transactions"]} | Missing Journal Entries: {report_data["missing_journal_count"]}'
    worksheet['A2'].font = Font(bold=True, color='DC2626' if report_data["missing_journal_count"] > 0 else '059669')
    worksheet['A2'].alignment = Alignment(horizontal='center')

    # Header row styling
    header_fill = PatternFill(start_color='D97706', end_color='D97706', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col_num in range(1, 8):
        cell = worksheet.cell(row=4, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Highlight missing journal entries
    alert_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    for row in range(5, len(audit_data_list) + 5):
        status_cell = worksheet.cell(row=row, column=7)
        if 'MISSING' in str(status_cell.value):
            for col in range(1, 8):
                worksheet.cell(row=row, column=col).fill = alert_fill

    # Column widths
    worksheet.column_dimensions['A'].width = 12
    worksheet.column_dimensions['B'].width = 20
    worksheet.column_dimensions['C'].width = 20
    worksheet.column_dimensions['D'].width = 30
    worksheet.column_dimensions['E'].width = 15
    worksheet.column_dimensions['F'].width = 20
    worksheet.column_dimensions['G'].width = 25

    writer.close()
    output.seek(0)

    filename = f'transaction_audit_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response = create_excel_response(filename)
    response.write(output.read())
    return response


def export_to_csv(data, columns, filename='export.csv'):
    """Generic CSV export function"""
    df = pd.DataFrame(data, columns=columns)

    response = create_csv_response(filename)
    df.to_csv(response, index=False)
    return response
